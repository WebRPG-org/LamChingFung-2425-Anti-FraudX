import os
import io
import tempfile
import logging
from typing import Optional, Dict, Any, List
import magic
import PyPDF2
from docx import Document
import openpyxl
import aiofiles
from fastapi import UploadFile, HTTPException
import re

logger = logging.getLogger(__name__)

class FileService:
    """檔案處理服務，支持多種文件格式的文字提取"""
    
    def __init__(self):
        self.supported_formats = {
            '.pdf': 'application/pdf',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.doc': 'application/msword',
            '.txt': 'text/plain',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.xls': 'application/vnd.ms-excel',
            '.csv': 'text/csv'
        }
        self.max_file_size = 10 * 1024 * 1024  # 10MB
        self.max_text_length = 50000  # 最大文字長度
    
    async def process_file(self, file: UploadFile) -> Dict[str, Any]:
        """
        處理上傳的文件並提取文字內容
        """
        try:
            # 驗證文件
            if not self._validate_file(file):
                raise HTTPException(status_code=400, detail="不支援的文件格式或文件過大")
            
            # 讀取文件內容
            content = await file.read()
            
            # 檢測文件類型
            file_type = self._detect_file_type(content, file.filename)
            
            # 提取文字內容
            text = await self._extract_text(content, file_type, file.filename)
            
            # 清理和處理文字
            processed_text = self._clean_text(text)
            
            return {
                "success": True,
                "text": processed_text,
                "file_info": {
                    "filename": file.filename,
                    "size": len(content),
                    "type": file_type,
                    "content_type": file.content_type,
                    "text_length": len(processed_text)
                }
            }
            
        except Exception as e:
            logger.error(f"File processing failed: {e}")
            return {
                "success": False,
                "error": f"文件處理失敗: {str(e)}",
                "text": ""
            }
    
    def _validate_file(self, file: UploadFile) -> bool:
        """驗證文件"""
        if not file.filename:
            return False
        
        # 檢查文件擴展名
        file_ext = os.path.splitext(file.filename.lower())[1]
        if file_ext not in self.supported_formats:
            return False
        
        return True
    
    def _detect_file_type(self, content: bytes, filename: str) -> str:
        """檢測文件類型"""
        try:
            # 使用python-magic檢測MIME類型
            mime_type = magic.from_buffer(content, mime=True)
            
            # 根據MIME類型和文件擴展名確定文件類型
            file_ext = os.path.splitext(filename.lower())[1]
            
            if mime_type == 'application/pdf' or file_ext == '.pdf':
                return 'pdf'
            elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 
                              'application/msword'] or file_ext in ['.docx', '.doc']:
                return 'docx'
            elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                              'application/vnd.ms-excel'] or file_ext in ['.xlsx', '.xls']:
                return 'xlsx'
            elif mime_type == 'text/plain' or file_ext == '.txt':
                return 'txt'
            elif mime_type == 'text/csv' or file_ext == '.csv':
                return 'csv'
            else:
                return 'unknown'
                
        except Exception as e:
            logger.warning(f"Failed to detect file type: {e}")
            # 回退到文件擴展名檢測
            file_ext = os.path.splitext(filename.lower())[1]
            if file_ext in self.supported_formats:
                return file_ext[1:]  # 移除點號
            return 'unknown'
    
    async def _extract_text(self, content: bytes, file_type: str, filename: str) -> str:
        """根據文件類型提取文字"""
        try:
            if file_type == 'pdf':
                return await self._extract_pdf_text(content)
            elif file_type == 'docx':
                return await self._extract_docx_text(content)
            elif file_type == 'xlsx':
                return await self._extract_xlsx_text(content)
            elif file_type in ['txt', 'csv']:
                return await self._extract_text_file(content)
            else:
                raise Exception(f"不支援的文件類型: {file_type}")
                
        except Exception as e:
            logger.error(f"Text extraction failed for {file_type}: {e}")
            raise Exception(f"文字提取失敗: {str(e)}")
    
    async def _extract_pdf_text(self, content: bytes) -> str:
        """提取PDF文字"""
        try:
            pdf_file = io.BytesIO(content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            text = ""
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                text += page.extract_text() + "\n"
                
                # 限制文字長度
                if len(text) > self.max_text_length:
                    text = text[:self.max_text_length] + "..."
                    break
            
            return text.strip()
            
        except Exception as e:
            logger.error(f"PDF text extraction failed: {e}")
            raise Exception(f"PDF文字提取失敗: {str(e)}")
    
    async def _extract_docx_text(self, content: bytes) -> str:
        """提取DOCX文字"""
        try:
            doc_file = io.BytesIO(content)
            doc = Document(doc_file)
            
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
                
                # 限制文字長度
                if len(text) > self.max_text_length:
                    text = text[:self.max_text_length] + "..."
                    break
            
            return text.strip()
            
        except Exception as e:
            logger.error(f"DOCX text extraction failed: {e}")
            raise Exception(f"DOCX文字提取失敗: {str(e)}")
    
    async def _extract_xlsx_text(self, content: bytes) -> str:
        """提取XLSX文字"""
        try:
            excel_file = io.BytesIO(content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"工作表: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                
                # 限制文字長度
                if len(text) > self.max_text_length:
                    text = text[:self.max_text_length] + "..."
                    break
            
            return text.strip()
            
        except Exception as e:
            logger.error(f"XLSX text extraction failed: {e}")
            raise Exception(f"XLSX文字提取失敗: {str(e)}")
    
    async def _extract_text_file(self, content: bytes) -> str:
        """提取純文字文件內容"""
        try:
            # 嘗試多種編碼
            encodings = ['utf-8', 'big5', 'gb2312', 'gbk', 'latin-1']
            
            for encoding in encodings:
                try:
                    text = content.decode(encoding)
                    return text.strip()
                except UnicodeDecodeError:
                    continue
            
            # 如果所有編碼都失敗，使用錯誤處理
            text = content.decode('utf-8', errors='ignore')
            return text.strip()
            
        except Exception as e:
            logger.error(f"Text file extraction failed: {e}")
            raise Exception(f"文字文件提取失敗: {str(e)}")
    
    def _clean_text(self, text: str) -> str:
        """清理和處理文字"""
        if not text:
            return ""
        
        # 移除多餘的空白字符
        text = re.sub(r'\s+', ' ', text)
        
        # 移除特殊字符但保留中文、英文、數字和基本標點
        text = re.sub(r'[^\u4e00-\u9fff\u3400-\u4dbf\w\s.,!?;:()（）【】「」""''、。，！？；：]', '', text)
        
        # 限制文字長度
        if len(text) > self.max_text_length:
            text = text[:self.max_text_length] + "..."
        
        return text.strip()
    
    def get_supported_formats(self) -> Dict[str, str]:
        """獲取支援的文件格式"""
        return self.supported_formats
    
    def get_max_file_size(self) -> int:
        """獲取最大文件大小"""
        return self.max_file_size
    
    def get_max_text_length(self) -> int:
        """獲取最大文字長度"""
        return self.max_text_length

# 創建全局實例
file_service = FileService()
